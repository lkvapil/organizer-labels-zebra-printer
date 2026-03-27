import sys
import json
import os
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QComboBox, QPushButton, 
                             QFileDialog, QMessageBox, QSpinBox, QGroupBox,
                             QTableWidget, QTableWidgetItem, QHeaderView)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont
from zebra import Zebra
import openpyxl
from datetime import datetime


class PrinterSelectorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.selected_printer = None
        self.file_path = None
        self.dpi = 203
        self.max_rows = 1
        self.label_width = 2 * 203  # 50mm = 2 inches at 203 DPI
        self.label_height = 1 * 203  # 25mm = 1 inch at 203 DPI
        self.print_rectangle = True
        self.rect_width_mm = 38
        self.rect_height_mm = 18
        self.config_file = os.path.join(os.path.dirname(__file__), 'printer_config.json')
        self.init_ui()
        
        # Automatically set path to organiser.xlsx
        default_file = os.path.join(os.path.dirname(__file__), 'organiser.xlsx')
        if os.path.exists(default_file):
            self.file_path = default_file
            self.file_label.setText(f"Selected file: {default_file}")
            self.load_preview(default_file)
            self.check_ready_to_print()
        
    def init_ui(self):
        self.setWindowTitle("Organizer labels with zebra printer")
        self.setGeometry(100, 100, 700, 750)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title = QLabel("Organizer labels with zebra printer")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)
        
        # Printer selection group
        printer_group = QGroupBox("Printer Selection")
        printer_layout = QVBoxLayout()
        
        # Label for printer
        printer_label = QLabel("Available printers:")
        printer_layout.addWidget(printer_label)
        
        # Combo box for printer selection
        self.printer_combo = QComboBox()
        self.printer_combo.setMinimumHeight(35)
        printer_layout.addWidget(self.printer_combo)
        
        # Refresh button
        refresh_btn = QPushButton("Refresh printer list")
        refresh_btn.clicked.connect(self.load_printers)
        refresh_btn.setMinimumHeight(35)
        printer_layout.addWidget(refresh_btn)
        
        printer_group.setLayout(printer_layout)
        main_layout.addWidget(printer_group)
        
        # Settings group
        settings_group = QGroupBox("Settings")
        settings_layout = QVBoxLayout()
        
        # Label size selection
        label_size_layout = QHBoxLayout()
        label_size_label = QLabel("Label size:")
        self.label_size_combo = QComboBox()
        self.label_size_combo.addItems([
            "50x25mm (2x1 inches)",
            "4x6 inches (102x152mm)",
            "4x3 inches (102x76mm)",
            "4x2 inches (102x51mm)",
            "3x2 inches (76x51mm)",
            "3x1 inches (76x25mm)",
            "2x1 inches (51x25mm)"
        ])
        self.label_size_combo.setCurrentText("50x25mm (2x1 inches)")
        self.label_size_combo.currentTextChanged.connect(self.update_label_size)
        label_size_layout.addWidget(label_size_label)
        label_size_layout.addWidget(self.label_size_combo)
        settings_layout.addLayout(label_size_layout)
        
        # DPI selection
        dpi_layout = QHBoxLayout()
        dpi_label = QLabel("DPI:")
        self.dpi_combo = QComboBox()
        self.dpi_combo.addItems(["203", "300"])
        self.dpi_combo.setCurrentText("203")
        self.dpi_combo.currentTextChanged.connect(self.update_dpi)
        dpi_layout.addWidget(dpi_label)
        dpi_layout.addWidget(self.dpi_combo)
        dpi_layout.addStretch()
        settings_layout.addLayout(dpi_layout)
        
        # Max rows per label
        rows_layout = QHBoxLayout()
        rows_label = QLabel("Max. labels:")
        self.rows_spin = QSpinBox()
        self.rows_spin.setMinimum(1)
        self.rows_spin.setMaximum(100)
        self.rows_spin.setValue(1)
        self.rows_spin.valueChanged.connect(self.update_max_rows)
        rows_layout.addWidget(rows_label)
        rows_layout.addWidget(self.rows_spin)
        rows_layout.addStretch()
        settings_layout.addLayout(rows_layout)
        
        # Rectangle settings
        from PyQt6.QtWidgets import QCheckBox
        self.rect_checkbox = QCheckBox("Print border rectangle around text")
        self.rect_checkbox.setChecked(True)
        self.rect_checkbox.stateChanged.connect(self.update_rectangle_enabled)
        settings_layout.addWidget(self.rect_checkbox)
        
        # Rectangle width
        rect_width_layout = QHBoxLayout()
        rect_width_label = QLabel("Border width (mm):")
        self.rect_width_spin = QSpinBox()
        self.rect_width_spin.setMinimum(10)
        self.rect_width_spin.setMaximum(100)
        self.rect_width_spin.setValue(40)
        self.rect_width_spin.valueChanged.connect(self.update_rect_width)
        rect_width_layout.addWidget(rect_width_label)
        rect_width_layout.addWidget(self.rect_width_spin)
        rect_width_layout.addStretch()
        settings_layout.addLayout(rect_width_layout)
        
        # Rectangle height
        rect_height_layout = QHBoxLayout()
        rect_height_label = QLabel("Border height (mm):")
        self.rect_height_spin = QSpinBox()
        self.rect_height_spin.setMinimum(10)
        self.rect_height_spin.setMaximum(100)
        self.rect_height_spin.setValue(18)
        self.rect_height_spin.valueChanged.connect(self.update_rect_height)
        rect_height_layout.addWidget(rect_height_label)
        rect_height_layout.addWidget(self.rect_height_spin)
        rect_height_layout.addStretch()
        settings_layout.addLayout(rect_height_layout)
        
        settings_group.setLayout(settings_layout)
        main_layout.addWidget(settings_group)
        
        # File selection group
        file_group = QGroupBox("Excel File")
        file_layout = QVBoxLayout()
        
        # File path display
        self.file_label = QLabel("No file selected")
        self.file_label.setWordWrap(True)
        file_layout.addWidget(self.file_label)
        
        # Select file button
        select_file_btn = QPushButton("Select Excel file")
        select_file_btn.clicked.connect(self.select_file)
        select_file_btn.setMinimumHeight(35)
        file_layout.addWidget(select_file_btn)

        # Excel preview table
        preview_label = QLabel("Preview (all rows):")
        file_layout.addWidget(preview_label)

        self.preview_table = QTableWidget()
        self.preview_table.setMinimumHeight(180)
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.preview_table.setAlternatingRowColors(True)
        file_layout.addWidget(self.preview_table)
        
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)
        
        # Print button
        self.print_btn = QPushButton("Print")
        self.print_btn.setMinimumHeight(45)
        self.print_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.print_btn.clicked.connect(self.print_label)
        self.print_btn.setEnabled(False)
        main_layout.addWidget(self.print_btn)
        
        # Status label
        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setWordWrap(True)
        main_layout.addWidget(self.status_label)
        
        main_layout.addStretch()
        
        # Load printers on startup
        self.load_printers()
        
    def load_printers(self):
        """Load all available printers from system"""
        try:
            import subprocess
            
            self.printer_combo.clear()
            all_printers = []
            
            print("=== Loading printers ===")
            
            # Method 1: Zebra library
            try:
                z = Zebra()
                zebra_printers = z.getqueues()
                print(f"Zebra library: {zebra_printers}")
                if zebra_printers:
                    all_printers.extend(zebra_printers)
            except Exception as e:
                print(f"Zebra library failed: {e}")
            
            # Method 2: lpstat -a (all available printers)
            try:
                result = subprocess.run(['lpstat', '-a'], 
                                      capture_output=True, 
                                      text=True, 
                                      timeout=5)
                print(f"lpstat -a output:\n{result.stdout}")
                if result.returncode == 0:
                    for line in result.stdout.split('\n'):
                        if line.strip():
                            printer_name = line.split()[0]
                            if printer_name not in all_printers:
                                all_printers.append(printer_name)
            except Exception as e:
                print(f"lpstat -a failed: {e}")
            
            # Method 3: lpstat -p (printer details)
            try:
                result = subprocess.run(['lpstat', '-p'], 
                                      capture_output=True, 
                                      text=True, 
                                      timeout=5)
                print(f"lpstat -p output:\n{result.stdout}")
                if result.returncode == 0:
                    for line in result.stdout.split('\n'):
                        if line.startswith('printer '):
                            printer_name = line.split()[1]
                            if printer_name not in all_printers:
                                all_printers.append(printer_name)
            except Exception as e:
                print(f"lpstat -p failed: {e}")
            
            # Method 4: lpstat -v (verbose)
            try:
                result = subprocess.run(['lpstat', '-v'], 
                                      capture_output=True, 
                                      text=True, 
                                      timeout=5)
                print(f"lpstat -v output:\n{result.stdout}")
                if result.returncode == 0:
                    for line in result.stdout.split('\n'):
                        if line.startswith('device for '):
                            printer_name = line.split('device for ')[1].split(':')[0]
                            if printer_name not in all_printers:
                                all_printers.append(printer_name)
            except Exception as e:
                print(f"lpstat -v failed: {e}")
            
            # Remove duplicates and sort
            all_printers = sorted(list(set(all_printers)))
            print(f"Total found: {all_printers}")
            
            if all_printers:
                self.printer_combo.addItems(all_printers)
                self.status_label.setText(f"Found {len(all_printers)} printer(s)")
                self.status_label.setStyleSheet("color: green;")
                
                # Load last used printer
                last_printer = self.load_last_printer()
                if last_printer and last_printer in all_printers:
                    self.printer_combo.setCurrentText(last_printer)
                    print(f"Restored last used printer: {last_printer}")
                
                self.check_ready_to_print()
            else:
                self.status_label.setText("No printers found! Check the terminal output.")
                self.status_label.setStyleSheet("color: red;")
                
        except Exception as e:
            print(f"Main error: {e}")
            self.status_label.setText(f"Error loading printers: {str(e)}")
            self.status_label.setStyleSheet("color: red;")
    
    def update_label_size(self, value):
        """Update label size based on selection"""
        label_sizes = {
            "50x25mm (2x1 inches)": (2 * 203, 1 * 203),
            "4x6 inches (102x152mm)": (4 * 203, 6 * 203),
            "4x3 inches (102x76mm)": (4 * 203, 3 * 203),
            "4x2 inches (102x51mm)": (4 * 203, 2 * 203),
            "3x2 inches (76x51mm)": (3 * 203, 2 * 203),
            "3x1 inches (76x25mm)": (3 * 203, 1 * 203),
            "2x1 inches (51x25mm)": (2 * 203, 1 * 203)
        }
        
        if value in label_sizes:
            self.label_width, self.label_height = label_sizes[value]
            print(f"Label size set: {value} -> {self.label_width}x{self.label_height} dots")
    
    def update_dpi(self, value):
        """Update DPI setting"""
        self.dpi = int(value)
        
    def update_max_rows(self, value):
        """Update max rows per label"""
        self.max_rows = value
    
    def update_rectangle_enabled(self, state):
        """Update rectangle printing"""
        self.print_rectangle = (state == 2)  # Qt.CheckState.Checked
    
    def update_rect_width(self, value):
        """Update rectangle width"""
        self.rect_width_mm = value
    
    def update_rect_height(self, value):
        """Update rectangle height"""
        self.rect_height_mm = value
    
    def select_file(self):
        """Open file dialog to select Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel file",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.file_path = file_path
            self.file_label.setText(f"Selected file: {file_path}")
            self.load_preview(file_path)
            self.check_ready_to_print()
    
    def load_preview(self, file_path):
        """Load and display first 10 rows of Excel file in the preview table"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            rows = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != '' for cell in row):
                    rows.append(row)

            if not rows:
                self.preview_table.setRowCount(0)
                self.preview_table.setColumnCount(0)
                return

            max_cols = max(len(r) for r in rows)
            self.preview_table.setRowCount(len(rows))
            self.preview_table.setColumnCount(max_cols)
            self.preview_table.setHorizontalHeaderLabels([f"Col {i+1}" for i in range(max_cols)])

            for r_idx, row in enumerate(rows):
                for c_idx in range(max_cols):
                    value = row[c_idx] if c_idx < len(row) else None
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.preview_table.setItem(r_idx, c_idx, item)
        except Exception as e:
            print(f"Preview error: {e}")

    def check_ready_to_print(self):
        """Check if all requirements are met for printing"""
        if self.file_path and self.printer_combo.currentText():
            self.print_btn.setEnabled(True)
        else:
            self.print_btn.setEnabled(False)
    
    def sanitize_zpl(self, zpl_content):
        """Sanitize ZPL content by replacing unsupported characters"""
        replacements = {
            "'": "'",
            "č": "c", "Č": "C", "ď": "d", "Ď": "D",
            "ě": "e", "Ě": "E", "ň": "n", "Ň": "N",
            "ř": "r", "Ř": "R", "š": "s", "Š": "S",
            "ť": "t", "Ť": "T", "ů": "u", "Ů": "U",
            "ý": "y", "Ý": "Y", "ž": "z", "Ž": "Z",
            "ø": "o", "Ø": "O", "å": "a", "Å": "A",
            "æ": "ae", "Æ": "AE", "ä": "a", "Ä": "A",
            "ö": "o", "Ö": "O", "ü": "u", "Ü": "U",
            "ß": "ss", "é": "e", "è": "e", "ê": "e",
            "ë": "e", "á": "a", "à": "a", "â": "a",
            "í": "i", "ì": "i", "î": "i", "ó": "o",
            "ò": "o", "ô": "o", "ú": "u", "ù": "u",
            "û": "u", "ñ": "n", "ç": "c",
        }
        
        for old_char, new_char in replacements.items():
            zpl_content = zpl_content.replace(old_char, new_char)
        
        zpl_content = zpl_content.encode('cp437', errors='replace').decode('cp437')
        return zpl_content
    
    def load_excel_data(self, file_path, sheet_name=None):
        """Load data from Excel sheet"""
        workbook = openpyxl.load_workbook(file_path)
        
        if sheet_name is None:
            sheet = workbook.active
        else:
            sheet = workbook[sheet_name]
            
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows
            if any(cell is not None and str(cell).strip() != '' for cell in row):
                data.append(row)
        
        print(f"Loaded {len(data)} rows from Excel (including header)")
        return data
    
    def generate_zpl(self, data):
        """Generate ZPL for label with DPI scaling"""
        scale = self.dpi / 203
        
        zpl = '^XA\n'
        zpl += f'^PW{int(self.label_width * scale)}\n^LL{int(self.label_height * scale)}\n'
        
        y_offset = int(50 * scale)
        x_coords = [int(v * scale) for v in [0, 35, 150, 280, 520]]
        
        font_height = int(24 * scale)
        font_width = int(24 * scale)
        row_spacing = int(30 * scale)
        
        # Add only data rows (skip header at data[0])
        for row in data[1:]:
            for i, col in enumerate(row):
                col_text = str(col) if col is not None else ""
                if i >= len(x_coords):
                    x_coords.append(x_coords[-1] + int(100 * scale))
                zpl += f'^FO{x_coords[i]},{y_offset}^A0N,{font_height},{font_width}^FD{col_text}^FS\n'
            y_offset += row_spacing
        
        zpl += '^XZ\n'
        return zpl
    
    def generate_zpl_single_label(self, row):
        """Generate ZPL for one label with centered text from one Excel row - each cell on separate line"""
        scale = self.dpi / 203
        
        zpl = '^XA\n'
        zpl += f'^PW{int(self.label_width * scale)}\n^LL{int(self.label_height * scale)}\n'
        
        # Get all non-empty cells and reverse the order (bottom to top)
        text_parts = [str(cell) for cell in row if cell is not None and str(cell).strip() != '']
        text_parts.reverse()  # Reverse order so first cell appears at top
        
        # Calculate positions for vertical centering of text
        font_height = int(40 * scale)
        font_width = int(40 * scale)
        line_spacing = int(50 * scale)
        
        # Calculate actual text height: (n-1) spacings + 1 font height
        if len(text_parts) > 0:
            total_text_height = (len(text_parts) - 1) * line_spacing + font_height
        else:
            total_text_height = 0
        
        # Center text block on label, then shift 0.5mm down
        text_offset_mm = 0.5  # Shift text 0.5mm downward
        text_start_y = int((self.label_height * scale - total_text_height) / 2 + (text_offset_mm * self.dpi / 25.4 * scale))
        
        # Debug output
        print(f"\n=== DEBUG ZPL ===")
        print(f"Label: {self.label_width}x{self.label_height} dots (scale={scale})")
        print(f"Label scaled: {int(self.label_width * scale)}x{int(self.label_height * scale)} dots")
        print(f"Font: {font_height}x{font_width}, spacing: {line_spacing}")
        print(f"Text parts: {len(text_parts)}, total height: {total_text_height}")
        print(f"Text start Y: {text_start_y} (with {text_offset_mm}mm offset)")
        print(f"=================\n")
        
        # Draw rectangle centered around original text position (before offset)
        if self.print_rectangle:
            # Rectangle dimensions in mm converted to dots
            rect_width = int(self.rect_width_mm * self.dpi / 25.4 * scale)
            rect_height = int(self.rect_height_mm * self.dpi / 25.4 * scale)
            rect_thickness = int(3 * scale)  # 3 dots thickness
            
            # Center rectangle on label (not following text offset)
            rect_x = int((self.label_width * scale - rect_width) / 2)
            rect_y = int((self.label_height * scale - rect_height) / 2)
            
            # Draw rectangle
            zpl += f'^FO{rect_x},{rect_y}^GB{rect_width},{rect_height},{rect_thickness}^FS\n'
        
        # Add each cell as a centered line
        y_offset = text_start_y
        for text in text_parts:
            # Use ^FB with width and center alignment (0,0 = left margin, C = center)
            zpl += f'^FO0,{y_offset}^A0N,{font_height},{font_width}^FB{int(self.label_width * scale)},1,0,C,0^FD{text}^FS\n'
            y_offset += line_spacing
        
        zpl += '^XZ\n'
        return zpl
    
    def generate_multiple_labels(self, data):
        """Generate multiple labels - one label per Excel row"""
        print(f"=== generate_multiple_labels ===")
        print(f"Total rows: {len(data)}")
        print(f"Max labels: {self.max_rows}")
        
        if len(data) == 0:
            print("Data is empty")
            return "^XA^XZ\n"
        
        # Take only max_rows (each row = one label)
        limited_rows = data[:self.max_rows]
        
        print(f"Printing {len(limited_rows)} label(s)")
        for i, row in enumerate(limited_rows):
            print(f"  Label {i+1}: {row}")
        print("================================")
        
        # Generate ZPL for all labels
        all_zpl = ""
        for row in limited_rows:
            all_zpl += self.generate_zpl_single_label(row)
        
        return all_zpl
    
    def print_label(self):
        """Print the label using selected printer"""
        if not self.file_path:
            QMessageBox.warning(self, "Error", "No file selected!")
            return
        
        if not self.printer_combo.currentText():
            QMessageBox.warning(self, "Error", "No printer selected!")
            return
        
        try:
            # Load Excel data
            self.status_label.setText("Loading data from Excel...")
            self.status_label.setStyleSheet("color: blue;")
            QApplication.processEvents()
            
            data = self.load_excel_data(self.file_path)
            
            if not data:
                QMessageBox.warning(self, "Error", "Excel file is empty!")
                return
            
            print(f"=== DEBUG: Loaded Excel data ===")
            print(f"Total rows: {len(data)}")
            for idx, row in enumerate(data):
                print(f"Row {idx}: {row}")
            print(f"================================")
            
            # Generate ZPL
            self.status_label.setText("Generating ZPL code...")
            QApplication.processEvents()
            
            # Calculate actual labels to print (each row = one label)
            actual_labels_to_print = min(len(data), self.max_rows)
            
            zpl_content = self.generate_multiple_labels(data)
            zpl_content = self.sanitize_zpl(zpl_content)
            
            # Print
            self.status_label.setText("Printing...")
            QApplication.processEvents()
            
            selected_printer = self.printer_combo.currentText()
            z = Zebra(selected_printer)
            z.output(zpl_content)
            
            # Save last used printer
            self.save_last_printer(selected_printer)
            
            self.status_label.setText(f"✓ Print complete! ({actual_labels_to_print} label(s))")
            self.status_label.setStyleSheet("color: green;")
            
            QMessageBox.information(
                self,
                "Success",
                f"Print job sent successfully to:\n{selected_printer}\n\n"
                f"Printed {actual_labels_to_print} label(s) (max. {self.max_rows})."
            )
            
        except Exception as e:
            self.status_label.setText(f"Print error!")
            self.status_label.setStyleSheet("color: red;")
            QMessageBox.critical(self, "Error", f"An error occurred while printing:\n{str(e)}")
    
    def save_last_printer(self, printer_name):
        """Save last used printer to config file"""
        try:
            config = {'last_printer': printer_name}
            with open(self.config_file, 'w') as f:
                json.dump(config, f)
            print(f"Saved printer: {printer_name}")
        except Exception as e:
            print(f"Failed to save printer: {e}")
    
    def load_last_printer(self):
        """Load last used printer from config file"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                return config.get('last_printer')
        except Exception as e:
            print(f"Failed to load printer: {e}")
        return None


def main():
    app = QApplication(sys.argv)
    window = PrinterSelectorGUI()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
