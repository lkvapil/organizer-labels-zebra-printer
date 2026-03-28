import sys
import json
import os
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QGridLayout, QLabel, QComboBox, QPushButton,
                             QFileDialog, QMessageBox, QSpinBox, QDoubleSpinBox,
                             QGroupBox, QTableWidget, QTableWidgetItem,
                             QHeaderView, QScrollArea, QCheckBox)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QPixmap
import urllib.request
import urllib.error
from zebra import Zebra
import openpyxl
from datetime import datetime


class PrinterSelectorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.selected_printer = None
        self.file_path = None
        self.dpi = 203
        self.max_rows = 1
        self.label_width = 2 * 203  # 50mm = 2 inches at 203 DPI
        self.label_height = 1 * 203  # 25mm = 1 inch at 203 DPI
        self.print_rectangle = True
        self.rect_width_mm = 38
        self.rect_height_mm = 18
        self.print_columns = 3  # default: first 3 columns
        self.norm_column_index = 5  # default: column 5 (1-based)
        self.font_size_override = 39  # default font size in dots
        self.norm_y_offset_mm = 4.0
        self.norm_x_offset_mm = 7.5
        self.norm_font_height = 25
        self.norm_font_width = 25
        self.coords_file = os.path.join(os.path.dirname(__file__), 'label_coords.json')
        self.config_file = os.path.join(os.path.dirname(__file__), 'printer_config.json')
        # Load saved coords from JSON into instance attrs
        self._apply_coords_from_file()
        self.init_ui()
        
        # Automatically set path to organiser.xlsx
        default_file = os.path.join(os.path.dirname(__file__), 'organiser.xlsx')
        if os.path.exists(default_file):
            self.file_path = default_file
            self.file_label.setText(f"Selected file: {default_file}")
            self.load_preview(default_file)
            self.check_ready_to_print()
        
    def init_ui(self):
        self.setWindowTitle("Organizer labels with zebra printer")
        self.setGeometry(100, 100, 800, 1230)

        # Central widget with scroll area so all settings are always accessible
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        outer_layout = QVBoxLayout(central_widget)
        outer_layout.setContentsMargins(0, 0, 0, 0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        outer_layout.addWidget(scroll_area)

        scroll_content = QWidget()
        scroll_area.setWidget(scroll_content)
        main_layout = QVBoxLayout(scroll_content)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title = QLabel("Organizer labels with zebra printer")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)
        
        # Printer selection group
        printer_group = QGroupBox("Printer Selection")
        printer_layout = QVBoxLayout()
        
        # Label for printer
        printer_label = QLabel("Available printers:")
        printer_layout.addWidget(printer_label)
        
        # Combo box for printer selection
        self.printer_combo = QComboBox()
        self.printer_combo.setMinimumHeight(35)
        printer_layout.addWidget(self.printer_combo)
        
        # Refresh button
        refresh_btn = QPushButton("Refresh printer list")
        refresh_btn.clicked.connect(self.load_printers)
        refresh_btn.setMinimumHeight(35)
        printer_layout.addWidget(refresh_btn)
        
        printer_group.setLayout(printer_layout)
        main_layout.addWidget(printer_group)
        
        # Settings group
        settings_group = QGroupBox("Settings")
        settings_layout = QGridLayout()
        settings_layout.setHorizontalSpacing(12)
        settings_layout.setVerticalSpacing(8)
        settings_layout.setColumnStretch(1, 1)
        settings_layout.setColumnStretch(3, 1)
        _r = 0

        # Label size (full width)
        label_size_label = QLabel("Label size:")
        self.label_size_combo = QComboBox()
        self.label_size_combo.addItems([
            "2x1 inches (50x25mm)",
            "4x6 inches (102x152mm)",
            "4x3 inches (102x76mm)",
            "4x2 inches (102x51mm)",
            "3x2 inches (76x51mm)",
            "3x1 inches (76x25mm)",
            "2x1 inches (51x25mm)"
        ])
        self.label_size_combo.setCurrentText("2x1 inches (50x25mm)")
        self.label_size_combo.currentTextChanged.connect(self.update_label_size)
        settings_layout.addWidget(label_size_label, _r, 0)
        settings_layout.addWidget(self.label_size_combo, _r, 1, 1, 3)
        _r += 1

        # DPI | Max. labels
        dpi_label = QLabel("DPI:")
        self.dpi_combo = QComboBox()
        self.dpi_combo.addItems(["203", "300"])
        self.dpi_combo.setCurrentText("203")
        self.dpi_combo.currentTextChanged.connect(self.update_dpi)
        settings_layout.addWidget(dpi_label, _r, 0)
        settings_layout.addWidget(self.dpi_combo, _r, 1)

        rows_label = QLabel("Max. labels:")
        self.rows_spin = QSpinBox()
        self.rows_spin.setMinimum(1)
        self.rows_spin.setMaximum(100)
        self.rows_spin.setValue(1)
        self.rows_spin.setMinimumWidth(90)
        self.rows_spin.valueChanged.connect(self.update_max_rows)
        settings_layout.addWidget(rows_label, _r, 2)
        settings_layout.addWidget(self.rows_spin, _r, 3)
        _r += 1

        # Print columns | Font size
        cols_label = QLabel("Print columns:")
        self.cols_spin = QSpinBox()
        self.cols_spin.setMinimum(0)
        self.cols_spin.setMaximum(100)
        self.cols_spin.setValue(3)
        self.cols_spin.setSpecialValueText("All")
        self.cols_spin.setMinimumWidth(90)
        self.cols_spin.valueChanged.connect(self.update_print_columns)
        settings_layout.addWidget(cols_label, _r, 0)
        settings_layout.addWidget(self.cols_spin, _r, 1)

        fsize_label = QLabel("Font size (dots):")
        self.fsize_spin = QSpinBox()
        self.fsize_spin.setMinimum(0)
        self.fsize_spin.setMaximum(300)
        self.fsize_spin.setValue(39)
        self.fsize_spin.setSpecialValueText("Auto")
        self.fsize_spin.setMinimumWidth(90)
        self.fsize_spin.setToolTip("Override automatic font size (dots). 0 = auto-calculated from label height.")
        self.fsize_spin.valueChanged.connect(self.update_font_size)
        settings_layout.addWidget(fsize_label, _r, 2)
        settings_layout.addWidget(self.fsize_spin, _r, 3)
        _r += 1

        # Norm name column (full width)
        norm_label = QLabel("Norm name column:")
        self.norm_spin = QSpinBox()
        self.norm_spin.setMinimum(0)
        self.norm_spin.setMaximum(100)
        self.norm_spin.setValue(5)
        self.norm_spin.setSpecialValueText("Off")
        self.norm_spin.setMinimumWidth(90)
        self.norm_spin.setToolTip("Column index (1-based) containing the norm name (e.g. DIN 912).\nPrinted in the lower right corner of the label.")
        self.norm_spin.valueChanged.connect(self.update_norm_column)
        settings_layout.addWidget(norm_label, _r, 0, 1, 2)
        settings_layout.addWidget(self.norm_spin, _r, 2, 1, 2)
        _r += 1

        # Norm name fine-tuning group
        norm_group = QGroupBox("Norm name position & font")
        norm_group_layout = QGridLayout()
        norm_group_layout.setHorizontalSpacing(12)
        norm_group_layout.setVerticalSpacing(8)
        norm_group_layout.setColumnStretch(1, 1)
        norm_group_layout.setColumnStretch(3, 1)

        # Y offset | Right margin
        norm_y_label = QLabel("Y offset from bottom (mm):")
        self.norm_y_spin = QDoubleSpinBox()
        self.norm_y_spin.setMinimum(-20.0)
        self.norm_y_spin.setMaximum(50.0)
        self.norm_y_spin.setSingleStep(0.5)
        self.norm_y_spin.setDecimals(1)
        self.norm_y_spin.setValue(self.norm_y_offset_mm)
        self.norm_y_spin.setMinimumWidth(90)
        self.norm_y_spin.setToolTip("Distance from label bottom edge to norm text top (mm). Increase to move text higher.")
        self.norm_y_spin.valueChanged.connect(self.update_norm_y_offset)
        norm_group_layout.addWidget(norm_y_label, 0, 0)
        norm_group_layout.addWidget(self.norm_y_spin, 0, 1)

        norm_x_label = QLabel("Right margin (mm):")
        self.norm_x_spin = QDoubleSpinBox()
        self.norm_x_spin.setMinimum(0.0)
        self.norm_x_spin.setMaximum(50.0)
        self.norm_x_spin.setSingleStep(0.5)
        self.norm_x_spin.setDecimals(1)
        self.norm_x_spin.setValue(self.norm_x_offset_mm)
        self.norm_x_spin.setMinimumWidth(90)
        self.norm_x_spin.setToolTip("Distance from the right edge of the label (mm). 0 = flush right. Increase to move text left.")
        self.norm_x_spin.valueChanged.connect(self.update_norm_x_offset)
        norm_group_layout.addWidget(norm_x_label, 0, 2)
        norm_group_layout.addWidget(self.norm_x_spin, 0, 3)

        # Font height | Font width
        norm_fh_label = QLabel("Font height (dots):")
        self.norm_fh_spin = QSpinBox()
        self.norm_fh_spin.setMinimum(6)
        self.norm_fh_spin.setMaximum(200)
        self.norm_fh_spin.setValue(self.norm_font_height)
        self.norm_fh_spin.setMinimumWidth(90)
        self.norm_fh_spin.valueChanged.connect(self.update_norm_font_height)
        norm_group_layout.addWidget(norm_fh_label, 1, 0)
        norm_group_layout.addWidget(self.norm_fh_spin, 1, 1)

        norm_fw_label = QLabel("Font width (dots):")
        self.norm_fw_spin = QSpinBox()
        self.norm_fw_spin.setMinimum(6)
        self.norm_fw_spin.setMaximum(200)
        self.norm_fw_spin.setValue(self.norm_font_width)
        self.norm_fw_spin.setMinimumWidth(90)
        self.norm_fw_spin.valueChanged.connect(self.update_norm_font_width)
        norm_group_layout.addWidget(norm_fw_label, 1, 2)
        norm_group_layout.addWidget(self.norm_fw_spin, 1, 3)

        norm_group.setLayout(norm_group_layout)
        settings_layout.addWidget(norm_group, _r, 0, 1, 4)
        _r += 1

        # Rectangle settings
        self.rect_checkbox = QCheckBox("Print border rectangle around text")
        self.rect_checkbox.setChecked(True)
        self.rect_checkbox.stateChanged.connect(self.update_rectangle_enabled)
        settings_layout.addWidget(self.rect_checkbox, _r, 0, 1, 4)
        _r += 1

        # Border width | Border height
        rect_width_label = QLabel("Border width (mm):")
        self.rect_width_spin = QSpinBox()
        self.rect_width_spin.setMinimum(10)
        self.rect_width_spin.setMaximum(100)
        self.rect_width_spin.setValue(40)
        self.rect_width_spin.setMinimumWidth(90)
        self.rect_width_spin.valueChanged.connect(self.update_rect_width)
        settings_layout.addWidget(rect_width_label, _r, 0)
        settings_layout.addWidget(self.rect_width_spin, _r, 1)

        rect_height_label = QLabel("Border height (mm):")
        self.rect_height_spin = QSpinBox()
        self.rect_height_spin.setMinimum(10)
        self.rect_height_spin.setMaximum(100)
        self.rect_height_spin.setValue(18)
        self.rect_height_spin.setMinimumWidth(90)
        self.rect_height_spin.valueChanged.connect(self.update_rect_height)
        settings_layout.addWidget(rect_height_label, _r, 2)
        settings_layout.addWidget(self.rect_height_spin, _r, 3)

        settings_group.setLayout(settings_layout)
        main_layout.addWidget(settings_group)
        
        # File selection group
        file_group = QGroupBox("Excel File")
        file_layout = QVBoxLayout()
        
        # File path display
        self.file_label = QLabel("No file selected")
        self.file_label.setWordWrap(True)
        file_layout.addWidget(self.file_label)
        
        # Select file button
        select_file_btn = QPushButton("Select Excel file")
        select_file_btn.clicked.connect(self.select_file)
        select_file_btn.setMinimumHeight(35)
        file_layout.addWidget(select_file_btn)

        # Excel preview table
        preview_label = QLabel("Preview (all rows):")
        file_layout.addWidget(preview_label)

        self.preview_table = QTableWidget()
        self.preview_table.setMinimumHeight(180)
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.preview_table.itemSelectionChanged.connect(self.preview_zpl_label)
        file_layout.addWidget(self.preview_table)

        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)

        # ZPL label preview group
        zpl_group = QGroupBox("Label Preview (Labelary)")
        zpl_layout = QVBoxLayout()

        self.zpl_preview_img = QLabel("Select a row to preview the label")
        self.zpl_preview_img.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.zpl_preview_img.setMinimumHeight(200)
        self.zpl_preview_img.setStyleSheet("border: 1px solid #ccc; background: #f9f9f9;")
        zpl_layout.addWidget(self.zpl_preview_img)

        self.zpl_status_label = QLabel("")
        self.zpl_status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        zpl_layout.addWidget(self.zpl_status_label)

        zpl_group.setLayout(zpl_layout)
        main_layout.addWidget(zpl_group)
        
        # Print button
        self.print_btn = QPushButton("Print")
        self.print_btn.setMinimumHeight(45)
        self.print_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.print_btn.clicked.connect(self.print_label)
        self.print_btn.setEnabled(False)
        main_layout.addWidget(self.print_btn)
        
        # Status label
        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setWordWrap(True)
        main_layout.addWidget(self.status_label)
        
        main_layout.addStretch()
        
        # Load printers on startup
        self.load_printers()
        
    def load_printers(self):
        """Load all available printers from system"""
        try:
            import subprocess
            
            self.printer_combo.clear()
            all_printers = []
            
            print("=== Loading printers ===")
            
            # Method 1: Zebra library
            try:
                z = Zebra()
                zebra_printers = z.getqueues()
                print(f"Zebra library: {zebra_printers}")
                if zebra_printers:
                    all_printers.extend(zebra_printers)
            except Exception as e:
                print(f"Zebra library failed: {e}")
            
            # Method 2: lpstat -a (all available printers)
            try:
                result = subprocess.run(['lpstat', '-a'], 
                                      capture_output=True, 
                                      text=True, 
                                      timeout=5)
                print(f"lpstat -a output:\n{result.stdout}")
                if result.returncode == 0:
                    for line in result.stdout.split('\n'):
                        if line.strip():
                            printer_name = line.split()[0]
                            if printer_name not in all_printers:
                                all_printers.append(printer_name)
            except Exception as e:
                print(f"lpstat -a failed: {e}")
            
            # Method 3: lpstat -p (printer details)
            try:
                result = subprocess.run(['lpstat', '-p'], 
                                      capture_output=True, 
                                      text=True, 
                                      timeout=5)
                print(f"lpstat -p output:\n{result.stdout}")
                if result.returncode == 0:
                    for line in result.stdout.split('\n'):
                        if line.startswith('printer '):
                            printer_name = line.split()[1]
                            if printer_name not in all_printers:
                                all_printers.append(printer_name)
            except Exception as e:
                print(f"lpstat -p failed: {e}")
            
            # Method 4: lpstat -v (verbose)
            try:
                result = subprocess.run(['lpstat', '-v'], 
                                      capture_output=True, 
                                      text=True, 
                                      timeout=5)
                print(f"lpstat -v output:\n{result.stdout}")
                if result.returncode == 0:
                    for line in result.stdout.split('\n'):
                        if line.startswith('device for '):
                            printer_name = line.split('device for ')[1].split(':')[0]
                            if printer_name not in all_printers:
                                all_printers.append(printer_name)
            except Exception as e:
                print(f"lpstat -v failed: {e}")
            
            # Remove duplicates and sort
            all_printers = sorted(list(set(all_printers)))
            print(f"Total found: {all_printers}")
            
            if all_printers:
                self.printer_combo.addItems(all_printers)
                self.status_label.setText(f"Found {len(all_printers)} printer(s)")
                self.status_label.setStyleSheet("color: green;")
                
                # Load last used printer
                last_printer = self.load_last_printer()
                if last_printer and last_printer in all_printers:
                    self.printer_combo.setCurrentText(last_printer)
                    print(f"Restored last used printer: {last_printer}")
                
                self.check_ready_to_print()
            else:
                self.status_label.setText("No printers found! Check the terminal output.")
                self.status_label.setStyleSheet("color: red;")
                
        except Exception as e:
            print(f"Main error: {e}")
            self.status_label.setText(f"Error loading printers: {str(e)}")
            self.status_label.setStyleSheet("color: red;")
    
    def update_label_size(self, value):
        """Update label size based on selection"""
        label_sizes = {
            "2x1 inches (50x25mm)": (2 * self.dpi, 1 * self.dpi),
            "4x6 inches (102x152mm)": (4 * self.dpi, 6 * self.dpi),
            "4x3 inches (102x76mm)": (4 * self.dpi, 3 * self.dpi),
            "4x2 inches (102x51mm)": (4 * self.dpi, 2 * self.dpi),
            "3x2 inches (76x51mm)": (3 * self.dpi, 2 * self.dpi),
            "3x1 inches (76x25mm)": (3 * self.dpi, 1 * self.dpi),
            "2x1 inches (51x25mm)": (2 * self.dpi, 1 * self.dpi)
        }

        if value in label_sizes:
            self.label_width, self.label_height = label_sizes[value]
            print(f"Label size set: {value} -> {self.label_width}x{self.label_height} dots")
            if hasattr(self, 'zpl_preview_img'):
                self.preview_zpl_label()

    def update_dpi(self, value):
        """Update DPI setting and recalculate label dimensions"""
        self.dpi = int(value)
        self.update_label_size(self.label_size_combo.currentText())

    def update_max_rows(self, value):
        """Update max rows per label"""
        self.max_rows = value

    def update_rectangle_enabled(self, state):
        """Update rectangle printing"""
        self.print_rectangle = (state == 2)  # Qt.CheckState.Checked
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def update_rect_width(self, value):
        """Update rectangle width"""
        self.rect_width_mm = value
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def update_rect_height(self, value):
        """Update rectangle height"""
        self.rect_height_mm = value
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def update_print_columns(self, value):
        """Update print columns limit (0 = all)"""
        self.print_columns = value
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def update_norm_column(self, value):
        """Update norm name column index (0 = off, 1-based)"""
        self.norm_column_index = value
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def update_font_size(self, value):
        """Update font size override (0 = auto)"""
        self.font_size_override = value
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def update_norm_y_offset(self, value):
        self.norm_y_offset_mm = value
        self.save_norm_coords()
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def update_norm_x_offset(self, value):
        self.norm_x_offset_mm = value
        self.save_norm_coords()
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def update_norm_font_height(self, value):
        self.norm_font_height = value
        self.save_norm_coords()
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def update_norm_font_width(self, value):
        self.norm_font_width = value
        self.save_norm_coords()
        if hasattr(self, 'zpl_preview_img'):
            self.preview_zpl_label()

    def save_norm_coords(self):
        """Save current norm coords to label_coords.json"""
        data = {
            "_comment": "Fine-tune label element positions.",
            "norm_name": {
                "_comment_y": "Distance from BOTTOM edge to norm text top (mm).",
                "y_offset_from_bottom_mm": self.norm_y_offset_mm,
                "x_offset_from_right_mm": self.norm_x_offset_mm,
                "_comment_font": "Font size in dots at 203 DPI. Scaled automatically for 300 DPI.",
                "font_height_dots": self.norm_font_height,
                "font_width_dots": self.norm_font_width
            }
        }
        try:
            with open(self.coords_file, 'w') as f:
                json.dump(data, f, indent=2)
            print(f"Saved coords to {self.coords_file}")
        except Exception as e:
            print(f"Failed to save coords: {e}")

    def _apply_coords_from_file(self):
        """Load label_coords.json and apply values to instance attrs (called once at startup)"""
        try:
            if os.path.exists(self.coords_file):
                with open(self.coords_file, 'r') as f:
                    data = json.load(f)
                nc = data.get("norm_name", {})
                self.norm_y_offset_mm = float(nc.get("y_offset_from_bottom_mm", self.norm_y_offset_mm))
                self.norm_x_offset_mm = float(nc.get("x_offset_from_right_mm", self.norm_x_offset_mm))
                self.norm_font_height = int(nc.get("font_height_dots", self.norm_font_height))
                self.norm_font_width = int(nc.get("font_width_dots", self.norm_font_width))
        except Exception as e:
            print(f"Coords file load error: {e}")

    def select_file(self):
        """Open file dialog to select Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel file",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.file_path = file_path
            self.file_label.setText(f"Selected file: {file_path}")
            self.load_preview(file_path)
            self.check_ready_to_print()
    
    def load_preview(self, file_path):
        """Load and display first 10 rows of Excel file in the preview table"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            rows = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != '' for cell in row):
                    rows.append(row)

            if not rows:
                self.preview_table.setRowCount(0)
                self.preview_table.setColumnCount(0)
                return

            max_cols = max(len(r) for r in rows)
            self.preview_table.setRowCount(len(rows))
            self.preview_table.setColumnCount(max_cols)
            self.preview_table.setHorizontalHeaderLabels([f"Col {i+1}" for i in range(max_cols)])

            for r_idx, row in enumerate(rows):
                for c_idx in range(max_cols):
                    value = row[c_idx] if c_idx < len(row) else None
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.preview_table.setItem(r_idx, c_idx, item)

            self.preview_table.selectRow(0)
        except Exception as e:
            print(f"Preview error: {e}")

    def preview_zpl_label(self):
        """Generate ZPL for selected row and fetch PNG preview from Labelary API"""
        selected = self.preview_table.selectedItems()
        if not selected:
            return

        row_idx = self.preview_table.currentRow()
        row_data = []
        for c in range(self.preview_table.columnCount()):
            item = self.preview_table.item(row_idx, c)
            row_data.append(item.text() if item and item.text() else None)

        # Carry forward norm text from previous rows if current cell is empty
        if self.norm_column_index > 0:
            norm_idx = self.norm_column_index - 1
            if norm_idx < len(row_data) and (row_data[norm_idx] is None or str(row_data[norm_idx]).strip() == ''):
                for prev_r in range(row_idx - 1, -1, -1):
                    item = self.preview_table.item(prev_r, norm_idx)
                    if item and item.text().strip():
                        row_data[norm_idx] = item.text().strip()
                        break

        try:
            zpl = self.generate_zpl_single_label(tuple(row_data))
            zpl = self.sanitize_zpl(zpl)

            dpmm = 8 if self.dpi == 203 else 12
            width_in = round(self.label_width / self.dpi, 2)
            height_in = round(self.label_height / self.dpi, 2)
            url = f"http://api.labelary.com/v1/printers/{dpmm}dpmm/labels/{width_in}x{height_in}/0/"

            self.zpl_status_label.setText("Fetching preview...")
            self.zpl_status_label.setStyleSheet("color: blue;")
            QApplication.processEvents()

            req = urllib.request.Request(url, data=zpl.encode('utf-8'), method='POST')
            req.add_header('Accept', 'image/png')
            with urllib.request.urlopen(req, timeout=10) as response:
                png_data = response.read()

            pixmap = QPixmap()
            pixmap.loadFromData(png_data)
            scaled = pixmap.scaled(400, 200, Qt.AspectRatioMode.KeepAspectRatio,
                                   Qt.TransformationMode.SmoothTransformation)
            self.zpl_preview_img.setPixmap(scaled)
            self.zpl_status_label.setText(f"Preview: row {row_idx + 1}")
            self.zpl_status_label.setStyleSheet("color: green;")

        except urllib.error.URLError:
            self.zpl_status_label.setText("Preview unavailable (no internet connection)")
            self.zpl_status_label.setStyleSheet("color: orange;")
        except Exception as e:
            self.zpl_status_label.setText(f"Preview error: {e}")
            self.zpl_status_label.setStyleSheet("color: red;")

    def check_ready_to_print(self):
        """Check if all requirements are met for printing"""
        if self.file_path and self.printer_combo.currentText():
            self.print_btn.setEnabled(True)
        else:
            self.print_btn.setEnabled(False)
    
    def sanitize_zpl(self, zpl_content):
        """Sanitize ZPL content by replacing unsupported characters"""
        replacements = {
            "'": "'",
            "č": "c", "Č": "C", "ď": "d", "Ď": "D",
            "ě": "e", "Ě": "E", "ň": "n", "Ň": "N",
            "ř": "r", "Ř": "R", "š": "s", "Š": "S",
            "ť": "t", "Ť": "T", "ů": "u", "Ů": "U",
            "ý": "y", "Ý": "Y", "ž": "z", "Ž": "Z",
            "ø": "o", "Ø": "O", "å": "a", "Å": "A",
            "æ": "ae", "Æ": "AE", "ä": "a", "Ä": "A",
            "ö": "o", "Ö": "O", "ü": "u", "Ü": "U",
            "ß": "ss", "é": "e", "è": "e", "ê": "e",
            "ë": "e", "á": "a", "à": "a", "â": "a",
            "í": "i", "ì": "i", "î": "i", "ó": "o",
            "ò": "o", "ô": "o", "ú": "u", "ù": "u",
            "û": "u", "ñ": "n", "ç": "c",
        }
        
        for old_char, new_char in replacements.items():
            zpl_content = zpl_content.replace(old_char, new_char)
        
        zpl_content = zpl_content.encode('cp437', errors='replace').decode('cp437')
        return zpl_content
    
    def load_excel_data(self, file_path, sheet_name=None):
        """Load data from Excel sheet"""
        workbook = openpyxl.load_workbook(file_path)
        
        if sheet_name is None:
            sheet = workbook.active
        else:
            sheet = workbook[sheet_name]
            
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows
            if any(cell is not None and str(cell).strip() != '' for cell in row):
                data.append(row)
        
        print(f"Loaded {len(data)} rows from Excel (including header)")
        return data
    
    def generate_zpl(self, data):
        """Generate ZPL for label with DPI scaling"""
        scale = self.dpi / 203
        
        zpl = '^XA\n'
        zpl += f'^PW{int(self.label_width * scale)}\n^LL{int(self.label_height * scale)}\n'
        
        y_offset = int(50 * scale)
        x_coords = [int(v * scale) for v in [0, 35, 150, 280, 520]]
        
        font_height = int(24 * scale)
        font_width = int(24 * scale)
        row_spacing = int(30 * scale)
        
        # Add only data rows (skip header at data[0])
        for row in data[1:]:
            for i, col in enumerate(row):
                col_text = str(col) if col is not None else ""
                if i >= len(x_coords):
                    x_coords.append(x_coords[-1] + int(100 * scale))
                zpl += f'^FO{x_coords[i]},{y_offset}^A0N,{font_height},{font_width}^FD{col_text}^FS\n'
            y_offset += row_spacing
        
        zpl += '^XZ\n'
        return zpl
    
    def generate_zpl_single_label(self, row):
        """Generate ZPL for one label with centered text from one Excel row - each cell on separate line"""
        # Extract norm name BEFORE slicing so it's independent of print_columns
        norm_text = None
        if self.norm_column_index > 0:
            norm_idx = self.norm_column_index - 1
            if norm_idx < len(row) and row[norm_idx] is not None and str(row[norm_idx]).strip():
                norm_text = str(row[norm_idx])

        if self.print_columns > 0:
            row = row[:self.print_columns]
        scale = self.dpi / 203
        
        zpl = '^XA\n'
        zpl += f'^PW{int(self.label_width * scale)}\n^LL{int(self.label_height * scale)}\n'
        
        # Get all non-empty cells and reverse the order (bottom to top)
        text_parts = [str(cell) for cell in row if cell is not None and str(cell).strip() != '']
        text_parts.reverse()  # Reverse order so first cell appears at top
        
        # Font scales proportionally with label height, adjusted for number of lines
        n_lines = max(len(text_parts), 1)
        if self.font_size_override > 0:
            font_height = int(self.font_size_override * scale)
        else:
            font_height = int(self.label_height * scale * 0.20 / (n_lines ** 0.5))
            font_height = max(20, min(font_height, 200))
        font_width = font_height
        line_spacing = int(font_height * 1.25)
        
        # Calculate actual text height: (n-1) spacings + 1 font height
        if len(text_parts) > 0:
            total_text_height = (len(text_parts) - 1) * line_spacing + font_height
        else:
            total_text_height = 0
        
        # Center text block on label, then shift 0.5mm down
        text_offset_mm = 0.5  # Shift text 0.5mm downward
        text_start_y = int((self.label_height * scale - total_text_height) / 2 + (text_offset_mm * self.dpi / 25.4 * scale))
        
        # Debug output
        print(f"\n=== DEBUG ZPL ===")
        print(f"Label: {self.label_width}x{self.label_height} dots (scale={scale})")
        print(f"Label scaled: {int(self.label_width * scale)}x{int(self.label_height * scale)} dots")
        print(f"Font: {font_height}x{font_width}, spacing: {line_spacing}")
        print(f"Text parts: {len(text_parts)}, total height: {total_text_height}")
        print(f"Text start Y: {text_start_y} (with {text_offset_mm}mm offset)")
        print(f"=================\n")
        
        # Draw rectangle centered around original text position (before offset)
        if self.print_rectangle:
            # Rectangle dimensions in mm converted to dots
            rect_width = int(self.rect_width_mm * self.dpi / 25.4 * scale)
            rect_height = int(self.rect_height_mm * self.dpi / 25.4 * scale)
            rect_thickness = int(3 * scale)  # 3 dots thickness
            
            # Center rectangle on label (not following text offset)
            rect_x = int((self.label_width * scale - rect_width) / 2)
            rect_y = int((self.label_height * scale - rect_height) / 2)
            
            # Draw rectangle
            zpl += f'^FO{rect_x},{rect_y}^GB{rect_width},{rect_height},{rect_thickness}^FS\n'
        
        # Add each cell as a centered line
        y_offset = text_start_y
        for text in text_parts:
            # Use ^FB with width and center alignment (0,0 = left margin, C = center)
            zpl += f'^FO0,{y_offset}^A0N,{font_height},{font_width}^FB{int(self.label_width * scale)},1,0,C,0^FD{text}^FS\n'
            y_offset += line_spacing

        # Draw norm name — always right-anchored; right margin shifts anchor left
        if norm_text:
            nfh = int(self.norm_font_height * scale)
            nfw = int(self.norm_font_width * scale)
            y_bottom_offset = int(self.norm_y_offset_mm * self.dpi / 25.4 * scale)
            right_margin_dots = int(self.norm_x_offset_mm * self.dpi / 25.4 * scale)
            norm_y = int(self.label_height * scale) - nfh - y_bottom_offset
            # anchor = right edge of text regardless of text length
            # 0 margin = flush to right edge; larger margin = further left
            anchor = max(10, int(self.label_width * scale) - right_margin_dots)
            zpl += f'^FO0,{norm_y}^A0N,{nfh},{nfw}^FB{anchor},1,0,R,0^FD{norm_text}^FS\n'

        zpl += '^XZ\n'
        return zpl
    
    def generate_multiple_labels(self, data):
        """Generate multiple labels - one label per Excel row"""
        print(f"=== generate_multiple_labels ===")
        print(f"Total rows: {len(data)}")
        print(f"Max labels: {self.max_rows}")
        
        if len(data) == 0:
            print("Data is empty")
            return "^XA^XZ\n"
        
        # Take only max_rows (each row = one label)
        limited_rows = data[:self.max_rows]
        
        print(f"Printing {len(limited_rows)} label(s)")
        for i, row in enumerate(limited_rows):
            print(f"  Label {i+1}: {row}")
        print("================================")
        
        # Generate ZPL for all labels — carry norm text forward from previous row if empty
        all_zpl = ""
        last_norm_text = None
        norm_idx = self.norm_column_index - 1  # 0-based
        for row in limited_rows:
            # If norm column is set and current cell is empty, inject last known norm value
            if self.norm_column_index > 0 and norm_idx < len(row):
                cell = row[norm_idx]
                if cell is None or str(cell).strip() == '':
                    if last_norm_text is not None:
                        row = list(row)
                        row[norm_idx] = last_norm_text
                        row = tuple(row)
                else:
                    last_norm_text = str(cell).strip()
            all_zpl += self.generate_zpl_single_label(row)

        return all_zpl
    
    def print_label(self):
        """Print the label using selected printer"""
        if not self.file_path:
            QMessageBox.warning(self, "Error", "No file selected!")
            return
        
        if not self.printer_combo.currentText():
            QMessageBox.warning(self, "Error", "No printer selected!")
            return
        
        try:
            # Load Excel data
            self.status_label.setText("Loading data from Excel...")
            self.status_label.setStyleSheet("color: blue;")
            QApplication.processEvents()
            
            data = self.load_excel_data(self.file_path)
            
            if not data:
                QMessageBox.warning(self, "Error", "Excel file is empty!")
                return
            
            print(f"=== DEBUG: Loaded Excel data ===")
            print(f"Total rows: {len(data)}")
            for idx, row in enumerate(data):
                print(f"Row {idx}: {row}")
            print(f"================================")
            
            # Generate ZPL
            self.status_label.setText("Generating ZPL code...")
            QApplication.processEvents()
            
            # Calculate actual labels to print (each row = one label)
            actual_labels_to_print = min(len(data), self.max_rows)
            
            zpl_content = self.generate_multiple_labels(data)
            zpl_content = self.sanitize_zpl(zpl_content)
            
            # Print
            self.status_label.setText("Printing...")
            QApplication.processEvents()
            
            selected_printer = self.printer_combo.currentText()
            z = Zebra(selected_printer)
            z.output(zpl_content)
            
            # Save last used printer
            self.save_last_printer(selected_printer)
            
            self.status_label.setText(f"✓ Print complete! ({actual_labels_to_print} label(s))")
            self.status_label.setStyleSheet("color: green;")
            
            QMessageBox.information(
                self,
                "Success",
                f"Print job sent successfully to:\n{selected_printer}\n\n"
                f"Printed {actual_labels_to_print} label(s) (max. {self.max_rows})."
            )
            
        except Exception as e:
            self.status_label.setText(f"Print error!")
            self.status_label.setStyleSheet("color: red;")
            QMessageBox.critical(self, "Error", f"An error occurred while printing:\n{str(e)}")
    
    def save_last_printer(self, printer_name):
        """Save last used printer to config file"""
        try:
            config = {'last_printer': printer_name}
            with open(self.config_file, 'w') as f:
                json.dump(config, f)
            print(f"Saved printer: {printer_name}")
        except Exception as e:
            print(f"Failed to save printer: {e}")
    
    def load_last_printer(self):
        """Load last used printer from config file"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                return config.get('last_printer')
        except Exception as e:
            print(f"Failed to load printer: {e}")
        return None


def main():
    app = QApplication(sys.argv)
    window = PrinterSelectorGUI()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
